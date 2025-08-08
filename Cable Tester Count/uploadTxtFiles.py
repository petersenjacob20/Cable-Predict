import re
from openpyxl import load_workbook

def parse_log_file(filepath):
    with open(filepath, 'r') as file:
        content = file.read()

    # Correct regex (not double-escaped)
    match_timestamp = re.search(r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\s+info\s+Test log started", content)
    if not match_timestamp:
        return None
    timestamp = match_timestamp.group(1)

    match_test_type = re.search(r"Test Type:\s*(\w+)", content)
    if not match_test_type:
        return None
    test_type = match_test_type.group(1).strip().upper()

    match_test_set = re.search(r"Test Set SN: (\d+)", content)
    match_coax = re.search(r"Coax Cable SN: (\d+)", content)
    match_signal = re.search(r"Signal Cable SN: (\d+)", content)

    if not (match_test_set and match_coax and match_signal):
        return None

    return {
        "timestamp": timestamp,
        "test_type": test_type,
        "test_set_sn": match_test_set.group(1),
        "coax_sn": match_coax.group(1),
        "signal_sn": match_signal.group(1)
    }

def is_already_logged(excel_file, timestamp):
    wb = load_workbook(excel_file)
    for sheet in wb.sheetnames:
        if sheet.startswith("CableTester Logs"):
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == timestamp:
                    wb.close()
                    return True
    wb.close()
    return False
