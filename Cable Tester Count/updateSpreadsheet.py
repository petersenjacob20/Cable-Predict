from openpyxl import load_workbook

def update_cable_counts(excel_file, data):
    wb = load_workbook(excel_file)
    test_type = data["test_type"].strip().upper()

    # Define sheet names per test type
    count_sheet_name = f"Cable Tester Count - {test_type}"
    log_sheet_name = f"CableTester Logs - {test_type}"

    # Ensure count sheet exists
    if count_sheet_name not in wb.sheetnames:
        ws_count = wb.create_sheet(count_sheet_name)
        ws_count.append(["Part Number", "Serial Number", "Usage Count"])
    else:
        ws_count = wb[count_sheet_name]

    def update_or_add_row(part_num, serial_num):
        for row in ws_count.iter_rows(min_row=2):
            if str(row[0].value) == part_num and str(row[1].value) == serial_num:
                row[2].value += 1
                return
        ws_count.append([part_num, serial_num, 1])

    # Assign part numbers dynamically based on test type
    update_or_add_row(f"{test_type} TESTER", data["test_set_sn"])
    update_or_add_row(f"{test_type} COAX CABLE", data["coax_sn"])
    update_or_add_row(f"{test_type} SIGNAL CABLE", data["signal_sn"])

    # Ensure log sheet exists
    if log_sheet_name not in wb.sheetnames:
        ws_log = wb.create_sheet(log_sheet_name)
        ws_log.append(["Timestamp", "Test Type", "Test Set SN", "Coax SN", "Signal SN"])
    else:
        ws_log = wb[log_sheet_name]

    ws_log.append([
        data["timestamp"],
        data["test_type"],
        data["test_set_sn"],
        data["coax_sn"],
        data["signal_sn"]
    ])

    wb.save(excel_file)
    wb.close()
