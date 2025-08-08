import argparse
from pathlib import Path

import pandas as pd
from lifelines import KaplanMeierFitter
from openpyxl import Workbook, load_workbook

# Default location for the workbook used by the CLI. The GUI can supply a
# different path when invoking these functions.
DEFAULT_EXCEL_FILE = (
    Path(__file__).resolve().parent / "Cable Tester Count" / "Cable Tracker.xlsx"
)


def add_event(
    excel_file: Path, connector_type: str, serial_number: str, cycles: int
) -> None:
    """Append a failure event to the Events sheet."""
    excel_file.parent.mkdir(parents=True, exist_ok=True)
    if excel_file.exists():
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
    if "Events" in wb.sheetnames:
        ws = wb["Events"]
    else:
        ws = wb.create_sheet("Events")
        ws.append(["connector_type", "serial_number", "cycles", "event"])
    ws.append([connector_type, serial_number, cycles, 1])
    wb.save(excel_file)


def analyze(excel_file: Path) -> None:
    """Perform survival analysis and write predictions."""
    if not excel_file.exists():
        print("Excel file not found.")
        return
    try:
        events_df = pd.read_excel(excel_file, sheet_name="Events")
    except ValueError:
        print("Events sheet not found.")
        return
    results = []
    for connector, group in events_df.groupby("connector_type"):
        kmf = KaplanMeierFitter()
        kmf.fit(group["cycles"], event_observed=group["event"])
        sf = kmf.survival_function_
        cycles80 = sf[sf["KM_estimate"] <= 0.8].index.min()
        cycles90 = sf[sf["KM_estimate"] <= 0.9].index.min()
        results.append(
            {
                "connector_type": connector,
                "median_cycles": kmf.median_survival_time_,
                "cycles_80_survival": float(cycles80) if pd.notna(cycles80) else None,
                "cycles_90_survival": float(cycles90) if pd.notna(cycles90) else None,
            }
        )
    pred_df = pd.DataFrame(results)
    wb = load_workbook(excel_file)
    if "Predictions" in wb.sheetnames:
        del wb["Predictions"]
    with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a") as writer:
        writer.book = wb
        writer.sheets = {ws.title: ws for ws in wb.worksheets}
        pred_df.to_excel(writer, sheet_name="Predictions", index=False)
    print("Analysis written to Predictions sheet.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Cable tracker CLI")
    subparsers = parser.add_subparsers(dest="command")

    add_parser = subparsers.add_parser("add-event", help="Add a cable event")
    add_parser.add_argument("--connector", required=True, help="Connector type")
    add_parser.add_argument("--serial", required=True, help="Serial number")
    add_parser.add_argument("--cycles", required=True, type=int, help="Cycle count")
    add_parser.add_argument(
        "--file",
        default=DEFAULT_EXCEL_FILE,
        type=Path,
        help="Path to Excel workbook",
    )

    analyze_parser = subparsers.add_parser("analyze", help="Run survival analysis")
    analyze_parser.add_argument(
        "--file",
        default=DEFAULT_EXCEL_FILE,
        type=Path,
        help="Path to Excel workbook",
    )

    args = parser.parse_args()
    if args.command == "add-event":
        add_event(args.file, args.connector, args.serial, args.cycles)
    elif args.command == "analyze":
        analyze(args.file)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
